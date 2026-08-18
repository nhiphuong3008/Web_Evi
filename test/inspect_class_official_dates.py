import os
import sys
import re
import datetime
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

official_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)"

def inspect_official_class_dates():
    class_start_dates = {}
    
    for root, dirs, files in os.walk(official_dir):
        excel_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$')]
        for f in excel_files:
            filepath = os.path.join(root, f)
            cname = ''
            # Regex match class code
            m = re.search(r'(Galax\s*\d+\.\d+|Sun\s*\d+\.\d+|Moon\s*\d+\.\d+|Sun\s*S\.\d+|S\.\d+)', f, re.IGNORECASE)
            if m:
                cname = m.group(1).strip().replace('S.7', 'Sun S.7')
            if not cname:
                m2 = re.search(r'(Galax|Sun|Moon)\s*\d+\.\d+', f, re.IGNORECASE)
                if m2: cname = m2.group(0).strip()
                
            if not cname:
                continue

            # Extract start date from filename if possible
            date_match = re.search(r'(\d{1,2})[_\/\-](\d{1,2})[_\/\-](\d{4})', f)
            parsed_date = None
            if date_match:
                d, month, y = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                try: parsed_date = datetime.date(y, month, d)
                except: pass

            # Inspect Excel DATES column
            first_excel_date = None
            excel_dates = []
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                for r in range(2, min(100, ws.max_row + 1)):
                    cell_v = ws.cell(row=r, column=1).value
                    if isinstance(cell_v, datetime.datetime) or isinstance(cell_v, datetime.date):
                        dt = cell_v.date() if isinstance(cell_v, datetime.datetime) else cell_v
                        excel_dates.append(dt)
                        if not first_excel_date:
                            first_excel_date = dt
            except Exception as e:
                pass

            effective_start = parsed_date or first_excel_date
            if cname not in class_start_dates or effective_start:
                class_start_dates[cname] = {
                    'class_name': cname,
                    'file': f,
                    'start_date': effective_start,
                    'dates_list': excel_dates
                }

    print(f"Discovered {len(class_start_dates)} official class date configs:\n")
    for cname in sorted(class_start_dates.keys()):
        cfg = class_start_dates[cname]
        s_date_str = cfg['start_date'].strftime('%Y-%m-%d (%A)') if cfg['start_date'] else 'None'
        clean_cname = cname.encode('ascii', 'ignore').decode('ascii')
        print(f"Class '{clean_cname}': StartDate = {s_date_str} (Has {len(cfg['dates_list'])} dates in sheet)")

if __name__ == '__main__':
    inspect_official_class_dates()
