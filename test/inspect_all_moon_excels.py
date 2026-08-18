import os
import sys
import glob
import json
import openpyxl

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

def parse_sheet_detail(ws):
    subtitle = ""
    title = ""
    vocab = []
    phonics = []
    struct = []

    current_cat = None
    for row in ws.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        row_str = [str(c).strip() if c is not None else "" for c in row]
        
        # Check titles
        full_line = " ".join(row_str).upper()
        if "MOON" in full_line and "UNIT TEST" in full_line:
            title = " ".join(row_str)
        elif "UNIT" in full_line and (":" in full_line or "LESSON" in full_line or "CLASSROOM" in full_line or "FAMILY" in full_line or "BODY" in full_line or "MY FACE" in full_line or "CLOTHES" in full_line or "HELLO" in full_line):
            if not subtitle:
                subtitle = " ".join(row_str)

        # Check Category markers
        col0 = row_str[0].lower()
        if "vocabulary" in col0 or "từ vựng" in col0:
            current_cat = "vocab"
        elif "phonics" in col0 or "ngữ âm" in col0:
            current_cat = "phonics"
        elif "mẫu câu" in col0 or "structure" in col0 or "sentences" in col0:
            current_cat = "struct"
        elif "comment" in col0 or "nhận xét" in col0:
            current_cat = "comment"

        # Content column is usually index 1 (or 0 if category span)
        content = ""
        if len(row_str) > 1 and row_str[1]:
            content = row_str[1]
        elif len(row_str) > 0 and row_str[0] and current_cat:
            c_val = row_str[0]
            if not any(k in c_val.lower() for k in ["catergory", "category", "vocabulary", "từ vựng", "phonics", "ngữ âm", "mẫu câu", "structure", "teacher comments", "nhận xét"]):
                content = c_val

        if content and current_cat:
            c_low = content.lower()
            if c_low in ["content", "nội dung", "excellent", "satisfactory", "need support", "catergory"]:
                continue
            if current_cat == "vocab":
                if content not in vocab: vocab.append(content)
            elif current_cat == "phonics":
                if content not in phonics: phonics.append(content)
            elif current_cat == "struct":
                if content not in struct: struct.append(content)

    return {
        "title": title,
        "subtitle": subtitle,
        "vocab": vocab,
        "phonics": phonics,
        "struct": struct
    }

def inspect_all():
    base_dir = "Moon test"
    all_data = {}
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                fpath = os.path.join(root, f)
                print("="*60)
                print(f"FILE: {fpath}")
                try:
                    wb = openpyxl.load_workbook(fpath, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        res = parse_sheet_detail(ws)
                        key = f"{os.path.basename(root)} :: {f} :: {sheet_name}"
                        all_data[key] = res
                        print(f"\n  [SHEET: {sheet_name}] Subtitle: {res['subtitle']}")
                        print(f"    - Vocab ({len(res['vocab'])}): {res['vocab']}")
                        print(f"    - Phonics ({len(res['phonics'])}): {res['phonics']}")
                        print(f"    - Struct ({len(res['struct'])}): {res['struct']}")
                except Exception as e:
                    print(f"    Error reading {fpath}: {e}")

    with open("test/full_moon_inspection.json", "w", encoding="utf-8") as out:
        json.dump(all_data, out, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    inspect_all()
