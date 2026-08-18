import os
import sys
import openpyxl

template_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\TEMPLATE"
galax_files = ["Galax 1.xlsx", "Galax 2.xlsx", "Galax 3.xlsx"]

for f in galax_files:
    filepath = os.path.join(template_dir, f)
    print(f"\n==========================================")
    print(f"FILE: {f}")
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n  Sheet '{sname}' ({ws.max_row} rows):")
            for r in range(1, min(12, ws.max_row + 1)):
                vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
                if any(v is not None for v in vals):
                    clean_vals = [str(v).replace('\n', ' ') if v is not None else '' for v in vals]
                    print(f"    Row {r}: {clean_vals[:6]}")
    except Exception as e:
        print(f"Error reading {f}: {e}")
