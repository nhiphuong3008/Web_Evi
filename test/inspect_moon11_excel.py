import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)\MT_B6_Moon 1.1 (2026).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Sheets in Moon 1.1 xlsx:", wb.sheetnames)
ws = wb.active

print("\nRows in Moon 1.1 syllabus:")
for row in ws.iter_rows(values_only=True):
    # Print row if not empty
    non_empty = [str(cell) for cell in row if cell is not None]
    if non_empty:
        print(non_empty[:6])
