import os
import openpyxl

template_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\TEMPLATE"
files = [f for f in os.listdir(template_dir) if f.endswith('.xlsx')]

print(f"Found {len(files)} excel files in TEMPLATE:")
for f in files:
    filepath = os.path.join(template_dir, f)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        print(f"\n--- File: {f} ---")
        print(f"Sheets: {sheet_names}")
        ws = wb.active
        print("Top 10 rows preview:")
        for row_idx in range(1, 11):
            row_vals = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 10)]
            if any(v is not None for v in row_vals):
                print(f"Row {row_idx}: {row_vals}")
    except Exception as e:
        print(f"Error reading {f}: {e}")
