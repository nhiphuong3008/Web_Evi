import os
import sys
import openpyxl

official_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)"

for root, dirs, files in os.walk(official_dir):
    excel_files = [f for f in files if f.endswith('.xlsx') and not f.startswith('~$')]
    for f in excel_files:
        filepath = os.path.join(root, f)
        relpath = os.path.relpath(filepath, official_dir)
        print(f"\n==========================================")
        print(f"FILE: {relpath}")
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            print(f"Sheets: {wb.sheetnames}")
            ws = wb.active
            print(f"Active Sheet '{ws.title}' ({ws.max_row} rows):")
            for r in range(1, min(10, ws.max_row + 1)):
                vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
                if any(v is not None for v in vals):
                    clean_vals = [str(v).replace('\n', ' ') if v is not None else '' for v in vals]
                    print(f"    Row {r}: {clean_vals[:6]}")
        except Exception as e:
            print(f"Error reading {f}: {e}")
