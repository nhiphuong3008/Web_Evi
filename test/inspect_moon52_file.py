import os
import sys
import openpyxl

filepath = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)\MT_B5_ Moon 5.2 (09_06_2026)_including Speaking Project.xlsx"

wb = openpyxl.load_workbook(filepath, data_only=True)
print(f"Sheets: {wb.sheetnames}")
ws = wb['Moon 5.2']

for r in range(1, 20):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
    clean_vals = [str(v).replace('\n', ' ') if v is not None else '' for v in vals]
    ascii_vals = [v.encode('ascii', 'ignore').decode('ascii') for v in clean_vals]
    print(f"Row {r:2d}: {ascii_vals[:8]}")
