import os
import sys
import openpyxl

official_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)"

for root, dirs, files in os.walk(official_dir):
    for f in files:
        if 'Galax 1.3' in f and f.endswith('.xlsx') and not f.startswith('~$'):
            filepath = os.path.join(root, f)
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sname in wb.sheetnames:
                if 'syllabus' in sname.lower():
                    ws = wb[sname]
                    clean_f = f.encode('ascii', 'ignore').decode('ascii')
                    print(f"\n==========================================")
                    print(f"File: {clean_f} | Sheet: {sname} ({ws.max_row} rows)")
                    for r in range(1, min(35, ws.max_row + 1)):
                        vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
                        ascii_vals = [str(v).replace('\n',' ').encode('ascii','ignore').decode('ascii') if v is not None else '' for v in vals]
                        if any(v for v in ascii_vals):
                            print(f"Row {r:2d}: {ascii_vals[:7]}")
