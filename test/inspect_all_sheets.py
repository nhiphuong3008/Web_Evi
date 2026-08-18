import openpyxl, os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

official_dir = r'c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)'
for f in os.listdir(official_dir):
    if f.endswith('.xlsx'):
        filepath = os.path.join(official_dir, f)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"=== File: {f} ===")
        print(f"    Sheets: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            # Find date count
            date_count = 0
            for r in rows:
                if r and r[0]:
                    s0 = str(r[0])
                    if '202' in s0 or '/' in s0:
                        date_count += 1
            print(f"    - Sheet '{sname}': {len(rows)} rows, ~{date_count} date entries")
