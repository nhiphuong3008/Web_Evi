import os
import sys
import glob
import json
import openpyxl

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

def parse_moon_excel(filepath):
    print(f"Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    print(f"  Sheets: {sheets}")
    
    parsed_units = {}
    for sheet_name in sheets:
        ws = wb[sheet_name]
        # Inspect title rows
        title = ""
        unit_subtitle = ""
        vocab_items = []
        phonics_items = []
        struct_items = []
        
        current_cat = None
        for row in ws.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            row_str = [str(c).strip() if c is not None else "" for c in row]
            
            # Check unit title
            for cell in row_str:
                if "UNIT" in cell.upper() and ("TEST" in cell.upper() or ":" in cell):
                    if not title: title = cell
                    elif not unit_subtitle: unit_subtitle = cell

            # Check categories
            first_cell = row_str[0].lower() if len(row_str) > 0 else ""
            second_cell = row_str[1] if len(row_str) > 1 else ""

            if "vocabulary" in first_cell or "từ vựng" in first_cell:
                current_cat = "vocab"
            elif "phonics" in first_cell or "ngữ âm" in first_cell:
                current_cat = "phonics"
            elif "mẫu câu" in first_cell or "structure" in first_cell:
                current_cat = "struct"
            elif "comment" in first_cell or "nhận xét" in first_cell:
                current_cat = "comment"

            content = second_cell if second_cell else (row_str[0] if current_cat and not first_cell.startswith("catergory") and not "vocabulary" in first_cell and not "phonics" in first_cell and not "mẫu câu" in first_cell else "")

            if content and current_cat:
                if content.lower() in ["content", "nội dung", "category", "catergory", "excellent", "satisfactory", "need support"]:
                    continue
                if current_cat == "vocab":
                    vocab_items.append(content)
                elif current_cat == "phonics":
                    phonics_items.append(content)
                elif current_cat == "struct":
                    struct_items.append(content)

        parsed_units[sheet_name] = {
            "title": title,
            "subtitle": unit_subtitle,
            "vocab": vocab_items,
            "phonics": phonics_items,
            "struct": struct_items
        }
    return parsed_units

def scan_all_moon_tests():
    moon_dir = "Moon test"
    results = {}
    for root, dirs, files in os.walk(moon_dir):
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                path = os.path.join(root, f)
                rel = os.path.relpath(path, moon_dir)
                try:
                    res = parse_moon_excel(path)
                    results[rel] = res
                except Exception as e:
                    print(f"Error reading {path}: {e}")
                    
    with open("test/moon_syllabus_db.json", "w", encoding="utf-8") as out:
        json.dump(results, out, ensure_ascii=False, indent=2)
    print("Done! Saved to test/moon_syllabus_db.json")

if __name__ == "__main__":
    scan_all_moon_tests()
