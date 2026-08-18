import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
filepath = r'c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)\TFB5_Sun 4.3_13_01_2026_including Speaking Project.xlsx'
wb = openpyxl.load_workbook(filepath, data_only=True)
print('Sheet names:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    print(f'=== Sheet: {sname} ({len(rows)} rows) ===')
    for idx, r in enumerate(rows):
        r_str = [str(c) if c is not None else '' for c in r]
        if any('50' in s for s in r_str) or any('7/8' in s for s in r_str) or any('ZOO' in s for s in r_str):
            vals = [str(c)[:30].replace('\n', ' ') if c is not None else '' for c in r[:9]]
            print(f' Row {idx:3d}: {vals}')
