import os
import openpyxl

template_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\TEMPLATE"
files = sorted([f for f in os.listdir(template_dir) if f.endswith('.xlsx')])

for f in files:
    filepath = os.path.join(template_dir, f)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"\n==========================================")
        print(f"FILE: {f}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_count = ws.max_row
            print(f"  Sheet '{sname}': {rows_count} rows")
            # print first 5 rows non-empty
            count = 0
            for r in range(1, min(15, rows_count + 1)):
                vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
                if any(v is not None for v in vals):
                    print(f"    R{r}: {vals[:6]}")
                    count += 1
                if count >= 6:
                    break
    except Exception as e:
        print(f"Error reading {f}: {e}")
