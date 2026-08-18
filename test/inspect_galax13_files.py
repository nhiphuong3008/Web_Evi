import os
import sys
import openpyxl

official_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)"

def check_galax13():
    files_matched = []
    for root, dirs, files in os.walk(official_dir):
        for f in files:
            if 'Galax 1.3' in f and f.endswith('.xlsx') and not f.startswith('~$'):
                files_matched.append(os.path.join(root, f))
                
    print(f"Found {len(files_matched)} files for Galax 1.3:")
    for fp in files_matched:
        fname = os.path.basename(fp)
        clean_f = fname.encode('ascii', 'ignore').decode('ascii')
        print(f"\n--- File: {clean_f} ---")
        wb = openpyxl.load_workbook(fp, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        ws = wb.active
        print(f"Active sheet: {ws.title} ({ws.max_row} rows)")
        for r in range(1, min(30, ws.max_row + 1)):
            v0 = ws.cell(row=r, column=1).value
            v1 = ws.cell(row=r, column=2).value
            v2 = ws.cell(row=r, column=3).value
            v3 = ws.cell(row=r, column=4).value
            v4 = ws.cell(row=r, column=5).value
            
            str_v0 = str(v0).strip() if v0 is not None else ''
            if '06/08' in str_v0 or str_v0 == '24' or str(v1) == '24':
                print(f"  Row {r:2d}: Date='{v0}' | Lesson='{v1}' | Unit='{v2}' | Pages='{v3}' | Vocab='{v4}'")

if __name__ == '__main__':
    check_galax13()
