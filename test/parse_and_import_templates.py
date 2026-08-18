import os
import sys
import re
import datetime
import openpyxl

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from database.db_manager import db_session, init_db
from database.models import LessonSyllabus
from services.db_service import detect_course_name_from_class

official_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\14. Class syllabus (official)"
template_dir = r"c:\Users\nhiph\OneDrive\Documents\Web_Evi\TEMPLATE"

def extract_class_name(filename):
    fname = os.path.basename(filename)
    m = re.search(r'(Galax\s*\d+\.\d+|Sun\s*\d+\.\d+|Moon\s*\d+\.\d+|Sun\s*S\.\d+|S\.\d+)', fname, re.IGNORECASE)
    cname = m.group(1).strip().replace('S.7', 'Sun S.7') if m else ''
    if not cname:
        m2 = re.search(r'(Galax|Sun|Moon)\s*\d+\.\d+', fname, re.IGNORECASE)
        if m2: cname = m2.group(0).strip()
    return cname

def parse_and_import_all_syllabuses():
    init_db()
    session = db_session()
    
    # Clean old syllabuses
    session.query(LessonSyllabus).delete()
    session.commit()

    total_inserted = 0

    # -------------------------------------------------------------
    # STEP 1: Parse Class-Specific Files from '14. Class syllabus (official)'
    # -------------------------------------------------------------
    class_files_map = {}
    for root, dirs, files in os.walk(official_dir):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                filepath = os.path.join(root, f)
                cname = extract_class_name(f)
                if cname:
                    if cname not in class_files_map:
                        class_files_map[cname] = []
                    class_files_map[cname].append((f, filepath))

    official_files = []
    for cname, flist in class_files_map.items():
        if cname == 'Galax 1.3':
            # Select exact file MT_B5_Galax 1.3 (20_4_2026).xlsx
            flist.sort(key=lambda x: 1 if 'doi gtrinh' not in x[0].lower() else 0, reverse=True)
        else:
            flist.sort(key=lambda x: (
                1 if ('doi gtrinh' in x[0].lower() or 'doi gtr' in x[0].lower() or 'gtrinh' in x[0].lower()) else 0,
                0 if 'finished' in x[0].lower() else 1,
                os.path.getmtime(x[1])
            ), reverse=True)
        best_file = flist[0]
        official_files.append((cname, best_file[0], best_file[1]))

    print(f"--- STEP 1: Importing {len(official_files)} Unique Class-Specific Syllabuses ---")
    for cname, f, filepath in official_files:
        course_name = detect_course_name_from_class(cname)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            # Find main syllabus sheet (Prioritize worksheet with actual date entries first!)
            target_sheet = None
            max_dates = 0
            
            for sname in wb.sheetnames:
                s = wb[sname]
                d_count = 0
                s_rows = list(s.iter_rows(values_only=True, max_row=120))
                for r in s_rows:
                    if r:
                        for cell in r[:2]:
                            if cell is not None:
                                c_str = str(cell)
                                if isinstance(cell, (datetime.datetime, datetime.date)) or ('/' in c_str and len(c_str) <= 10) or '202' in c_str:
                                    d_count += 1
                                    break
                if d_count > max_dates:
                    max_dates = d_count
                    target_sheet = s
            
            # If no sheet has dates, fallback to name priority (exact 'Syllabus', 'Basic', or cname)
            if not target_sheet or max_dates == 0:
                for sname in wb.sheetnames:
                    if sname.strip().lower() in ['syllabus', 'basic', cname.strip().lower()]:
                        target_sheet = wb[sname]
                        break
            if not target_sheet or max_dates == 0:
                for sname in wb.sheetnames:
                    if 'syllabus' in sname.lower() or 'basic' in sname.lower() or cname.lower() in sname.lower():
                        if 'gt4' not in sname.lower() and 'gt1' not in sname.lower() and 'old' not in sname.lower() and 'report' not in sname.lower():
                            target_sheet = wb[sname]
                            break
            if not target_sheet:
                target_sheet = wb.active

            ws = target_sheet
            rows = list(ws.iter_rows(values_only=True))
            
            header_idx = -1
            for idx, r in enumerate(rows[:15]):
                r_str = [str(cell).upper() if cell else '' for cell in r]
                if any('LESSON' in cell for cell in r_str) or any('UNIT' in cell for cell in r_str):
                    header_idx = idx
                    break

            if header_idx == -1: header_idx = 2

            lesson_count = 0
            for r_idx in range(header_idx + 1, len(rows)):
                r = rows[r_idx]
                if not r or not any(r): continue

                r_vals = [str(cell).strip() if cell is not None else '' for cell in r]
                raw_cell0 = r[0]

                # Extract date if present in column 1 (DATES)
                off_date_str = ''
                if isinstance(raw_cell0, (datetime.datetime, datetime.date)):
                    dt = raw_cell0.date() if isinstance(raw_cell0, datetime.datetime) else raw_cell0
                    off_date_str = dt.strftime('%Y-%m-%d')
                elif isinstance(raw_cell0, str):
                    raw_s = raw_cell0.strip()
                    m_full = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', raw_s)
                    m_short = re.search(r'(\d{1,2})/(\d{1,2})', raw_s)
                    if m_full:
                        d, m, y = int(m_full.group(1)), int(m_full.group(2)), int(m_full.group(3))
                        off_date_str = f"{y:04d}-{m:02d}-{d:02d}"
                    elif m_short:
                        d, m = int(m_short.group(1)), int(m_short.group(2))
                        off_date_str = f"2026-{m:02d}-{d:02d}"

                first_col = r_vals[0] if len(r_vals) > 0 else ''
                second_col = r_vals[1] if len(r_vals) > 1 else ''
                third_col = r_vals[2] if len(r_vals) > 2 else ''

                lesson_str = ''
                if 'LESSON' in second_col.upper():
                    lesson_str = second_col
                elif second_col.replace('.0','').isdigit():
                    lesson_str = f"LESSON {second_col.replace('.0','')}"
                elif 'LESSON' in first_col.upper():
                    lesson_str = first_col
                elif first_col.replace('.0','').isdigit() and not off_date_str:
                    lesson_str = f"LESSON {first_col.replace('.0','')}"

                if not lesson_str: continue

                unit_str = third_col if len(r_vals) > 2 else ''

                # Filter out header rows
                if lesson_str.upper() in ['LESSONS', 'LESSON', 'BUỔI'] and unit_str.upper() in ['UNITS', 'UNIT', 'NỘI DUNG']:
                    continue
                if unit_str.upper() in ['UNITS', 'UNIT', 'STUDENT\'S BOOK & ACTIVITY\'S BOOK PAGES', 'STUDENT\'S BOOK PAGES']:
                    continue

                lesson_count += 1
                pages_str = r_vals[3] if len(r_vals) > 3 else ''
                
                is_galax = 'galax' in course_name.lower() or 'galax' in cname.lower()
                if is_galax:
                    vocab = r_vals[4] if len(r_vals) > 4 else ''
                    grammar = r_vals[5] if len(r_vals) > 5 else ''
                    target = ''
                    hw_teacher = r_vals[6] if len(r_vals) > 6 else ''
                    hw_cm = r_vals[7] if len(r_vals) > 7 else ''
                else:
                    vocab = r_vals[4] if len(r_vals) > 4 else ''
                    grammar = r_vals[5] if len(r_vals) > 5 else ''
                    target = r_vals[6] if len(r_vals) > 6 else ''
                    hw_teacher = r_vals[7] if len(r_vals) > 7 else ''
                    hw_cm = r_vals[8] if len(r_vals) > 8 else ''

                syl = LessonSyllabus(
                    course_name=course_name,
                    class_name=cname,
                    official_date=off_date_str,
                    lesson_num=lesson_count,
                    lesson_title=f"LESSON {lesson_count}",
                    unit_name=unit_str,
                    pages=pages_str,
                    vocabulary=vocab,
                    grammar=grammar,
                    lesson_target=target,
                    homework_teacher=hw_teacher,
                    homework_cm=hw_cm,
                    file_source=f
                )
                session.add(syl)
                total_inserted += 1

            clean_cname = cname.encode('ascii', 'ignore').decode('ascii')
            print(f"  [CLASS SPECIFIC] '{clean_cname}' ===> Inserted {lesson_count} lessons")

        except Exception as e:
            print(f"Error parsing official file {f}: {e}")

    # -------------------------------------------------------------
    # STEP 2: Parse General Template Files from TEMPLATE folder
    # -------------------------------------------------------------
    template_files = sorted([f for f in os.listdir(template_dir) if f.endswith('.xlsx')])
    print(f"\n--- STEP 2: Importing {len(template_files)} General Template Syllabuses ---")
    
    for f in template_files:
        course_name = f.replace('.xlsx', '').strip()
        filepath = os.path.join(template_dir, f)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            target_sheet = None
            for sname in wb.sheetnames:
                if sname.lower() in [course_name.lower(), 'standard', 'syllabus', 'gt1 syllabus', 'gt2 syllabus', 'gt3 syllabus', 'gt4 syllabus', 'gt5 syllabus']:
                    target_sheet = wb[sname]
                    break
            if not target_sheet: target_sheet = wb.worksheets[0]

            ws = target_sheet
            rows = list(ws.iter_rows(values_only=True))

            header_idx = -1
            for idx, r in enumerate(rows[:15]):
                r_str = [str(cell).upper() if cell else '' for cell in r]
                if any('LESSON' in cell for cell in r_str) and any('UNIT' in cell for cell in r_str):
                    header_idx = idx
                    break

            if header_idx == -1: header_idx = 1

            lesson_count = 0
            for r_idx in range(header_idx + 1, len(rows)):
                r = rows[r_idx]
                if not r or not any(r): continue

                r_vals = [str(cell).strip() if cell is not None else '' for cell in r]
                first_col = r_vals[0] if len(r_vals) > 0 else ''
                second_col = r_vals[1] if len(r_vals) > 1 else ''
                third_col = r_vals[2] if len(r_vals) > 2 else ''

                lesson_str = ''
                if 'LESSON' in first_col.upper(): lesson_str = first_col
                elif 'LESSON' in second_col.upper(): lesson_str = second_col
                elif first_col.replace('.0','').isdigit(): lesson_str = f"LESSON {first_col.replace('.0','')}"
                elif second_col.replace('.0','').isdigit(): lesson_str = f"LESSON {second_col.replace('.0','')}"

                if not lesson_str: continue

                lesson_count += 1
                unit_str = third_col if len(r_vals) > 2 else ''
                pages_str = r_vals[3] if len(r_vals) > 3 else ''
                
                is_galax = 'galax' in course_name.lower()
                if is_galax:
                    vocab = r_vals[4] if len(r_vals) > 4 else ''
                    grammar = r_vals[5] if len(r_vals) > 5 else ''
                    target = ''
                    hw_teacher = r_vals[6] if len(r_vals) > 6 else ''
                    hw_cm = r_vals[7] if len(r_vals) > 7 else ''
                else:
                    vocab = r_vals[4] if len(r_vals) > 4 else ''
                    grammar = r_vals[5] if len(r_vals) > 5 else ''
                    target = r_vals[6] if len(r_vals) > 6 else ''
                    hw_teacher = r_vals[7] if len(r_vals) > 7 else ''
                    hw_cm = r_vals[8] if len(r_vals) > 8 else ''

                syl = LessonSyllabus(
                    course_name=course_name,
                    class_name=None, # General template
                    official_date='',
                    lesson_num=lesson_count,
                    lesson_title=lesson_str,
                    unit_name=unit_str,
                    pages=pages_str,
                    vocabulary=vocab,
                    grammar=grammar,
                    lesson_target=target,
                    homework_teacher=hw_teacher,
                    homework_cm=hw_cm,
                    file_source=f
                )
                session.add(syl)
                total_inserted += 1

            clean_course = course_name.encode('ascii', 'ignore').decode('ascii')
            print(f"  [GENERAL TEMPLATE] '{clean_course}' ===> Inserted {lesson_count} lessons")

        except Exception as e:
            print(f"Error parsing template {f}: {e}")

    session.commit()
    print(f"\n==========================================")
    print(f"SUCCESSFULLY IMPORTED {total_inserted} LESSON SYLLABUSES (CLASS-SPECIFIC & GENERAL) TO DATABASE!")
    session.close()

if __name__ == '__main__':
    parse_and_import_all_syllabuses()
